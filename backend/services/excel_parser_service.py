from __future__ import annotations

import json
import zipfile
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


class ExcelParserService:
    """Servicio para parsear archivos Excel de papeles de trabajo."""

    EXPECTED_COLUMNS = [
        "Tarea",
        "NIA",
        "Descripcion",
        "Evidencia",
        "Hallazgo",
        "Estado",
    ]

    @staticmethod
    def parse_excel(file_content: bytes) -> dict[str, Any]:
        """
        Parsea contenido de Excel (bytes) a diccionario JSON.

        Args:
            file_content: Contenido del archivo Excel en bytes

        Returns:
            Dict con datos parseados {rows: [...], summary: {...}, errors: [...]}
        """
        try:
            workbook = load_workbook(BytesIO(file_content), data_only=True)
            worksheet = workbook.active

            if not worksheet:
                return {"rows": [], "summary": {}, "errors": ["No worksheet found"]}

            rows = []
            errors = []

            # Leer headers (primera fila)
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())

            if not headers:
                return {"rows": [], "summary": {}, "errors": ["No headers found"]}

            # Leer datos (desde fila 2)
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue

                row_data = {}
                for col_idx, header in enumerate(headers):
                    value = row[col_idx] if col_idx < len(row) else None
                    row_data[header] = value if value is not None else ""

                # Validación básica
                if not row_data.get("Tarea"):
                    errors.append(f"Row {row_idx}: Tarea is required")
                    continue

                rows.append(row_data)

            summary = {
                "total_rows": len(rows),
                "completed": sum(1 for r in rows if str(r.get("Estado", "")).lower() == "completado"),
                "pending": sum(1 for r in rows if str(r.get("Estado", "")).lower() != "completado"),
                "columns": headers,
            }

            return {
                "rows": rows,
                "summary": summary,
                "errors": errors,
            }

        except Exception as e:
            return {
                "rows": [],
                "summary": {},
                "errors": [f"Error parsing Excel: {str(e)}"],
            }

    @staticmethod
    def compress_file(file_content: bytes, filename: str = "papeles.xlsx") -> bytes:
        """
        Comprime archivo Excel a ZIP.

        Args:
            file_content: Contenido del archivo original
            filename: Nombre del archivo a guardar en ZIP

        Returns:
            Contenido del archivo ZIP (bytes)
        """
        try:
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                zip_file.writestr(filename, file_content)
            return zip_buffer.getvalue()
        except Exception as e:
            raise ValueError(f"Error compressing file: {str(e)}")

    @staticmethod
    def decompress_file(zip_content: bytes) -> bytes:
        """
        Descomprime archivo ZIP.

        Args:
            zip_content: Contenido del archivo ZIP

        Returns:
            Contenido del primer archivo en el ZIP
        """
        try:
            zip_buffer = BytesIO(zip_content)
            with zipfile.ZipFile(zip_buffer, "r") as zip_file:
                files = zip_file.namelist()
                if not files:
                    raise ValueError("ZIP file is empty")
                return zip_file.read(files[0])
        except Exception as e:
            raise ValueError(f"Error decompressing file: {str(e)}")

    @staticmethod
    def create_template_excel() -> bytes:
        """
        Crea plantilla Excel vacía con estructura correcta.

        Returns:
            Contenido de archivo Excel (bytes)
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Papeles de Trabajo"

        # Headers
        headers = ExcelParserService.EXPECTED_COLUMNS
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="041627", end_color="041627", fill_type="solid")

        # Añadir ejemplo (fila 2)
        example = [
            "Ejemplo: Verificar saldos",
            "NIA 330",
            "Cotejo con TB vs Mayor",
            "Confirmación bancaria",
            "Sin diferencias",
            "Completado",
        ]
        for col_idx, value in enumerate(example, start=1):
            worksheet.cell(row=2, column=col_idx).value = value

        # Ajustar anchos de columna
        for col in worksheet.columns:
            worksheet.column_dimensions[col[0].column_letter].width = 20

        # Guardar en BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
