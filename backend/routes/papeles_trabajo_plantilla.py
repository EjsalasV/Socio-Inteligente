"""
Rutas para descargar plantillas de Papeles de Trabajo (V1 - Simple)
- Descargar plantilla de todos los papeles
- Descargar plantilla por Línea de Cuenta (L/S)
- Obtener papeles por L/S como JSON
"""
from typing import Any, Optional
from io import BytesIO

from fastapi import APIRouter, Depends, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func

from backend.auth import get_current_user
from backend.models.workpapers_template import WorkpapersTemplate
from backend.schemas import UserContext, ApiResponse
from backend.utils.database import get_session
from backend.utils.api_errors import raise_api_error

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    raise ImportError("openpyxl is required for Excel generation")

router = APIRouter(prefix="/api/papeles-trabajo", tags=["papeles-plantilla"])


@router.get("/{cliente_id}/papeles-count", response_model=ApiResponse)
async def contar_papeles(
    cliente_id: str,
    user: UserContext = Depends(get_current_user),
    session: Any = Depends(get_session),
) -> ApiResponse:
    """
    Contar papeles por L/S para debug
    """
    try:
        # Contar papeles por L/S
        papeles_count = (
            session.query(WorkpapersTemplate.ls)
            .group_by(WorkpapersTemplate.ls)
            .with_entities(WorkpapersTemplate.ls, func.count(WorkpapersTemplate.id))
            .all()
        )

        total = session.query(func.count(WorkpapersTemplate.id)).scalar()

        data = {
            "total": total,
            "por_ls": [{"ls": ls, "count": count} for ls, count in papeles_count],
        }

        return ApiResponse(status="success", data=data)

    except Exception as e:
        raise_api_error(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            code="ERROR",
            message=str(e),
        )


@router.get("/{cliente_id}/papeles-por-ls", response_model=ApiResponse)
async def obtener_papeles_por_ls(
    cliente_id: str,
    ls: Optional[int] = Query(None, description="Línea de Cuenta (ej: 130, 140)"),
    user: UserContext = Depends(get_current_user),
    session: Any = Depends(get_session),
) -> ApiResponse:
    """
    Obtener papeles de trabajo por Línea de Cuenta (L/S)

    Retorna lista de papeles con toda su información para visualización y gestión de observaciones.

    Parámetros:
    - cliente_id: ID del cliente
    - ls: Línea de Cuenta (ej: 130) - Requerido
    """
    try:
        if ls is None:
            raise_api_error(
                status_code=status.HTTP_400_BAD_REQUEST,
                code="LS_REQUIRED",
                message="El parámetro 'ls' (Línea de Cuenta) es requerido",
            )

        # Obtener papeles para la L/S
        papeles = (
            session.query(WorkpapersTemplate)
            .filter(WorkpapersTemplate.ls == ls)
            .order_by(WorkpapersTemplate.numero)
            .all()
        )

        if not papeles:
            raise_api_error(
                status_code=status.HTTP_404_NOT_FOUND,
                code="NO_PAPELES_FOUND",
                message=f"No se encontraron papeles de trabajo para L/S {ls}",
            )

        # Convertir a diccionarios
        papeles_data = [papel.to_dict() for papel in papeles]

        return ApiResponse(
            status="success",
            data={
                "ls": ls,
                "total": len(papeles_data),
                "papeles": papeles_data,
            },
        )

    except Exception as e:
        if hasattr(e, "status_code"):
            raise
        raise_api_error(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            code="ERROR_FETCHING_PAPELES",
            message=f"Error obteniendo papeles: {str(e)}",
        )


def _create_template_workbook(papeles: list[WorkpapersTemplate]) -> BytesIO:
    """Crear un workbook Excel con la estructura de papeles de trabajo"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Papeles de Trabajo"

    # Definir estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = [
        "Código",
        "L/S",
        "Nombre del Papel",
        "Aseveración",
        "Importancia",
        "Obligatorio",
        "Descripción/Motivo",
        "Observaciones",
        "Hallazgo",
        "Efecto Financiero",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Ajustar ancho de columnas
    column_widths = [12, 8, 30, 15, 12, 12, 25, 20, 25, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width

    # Agregar datos de papeles
    for row_num, papel in enumerate(papeles, 2):
        ws.cell(row=row_num, column=1).value = papel.codigo
        ws.cell(row=row_num, column=2).value = papel.ls
        ws.cell(row=row_num, column=3).value = papel.nombre
        ws.cell(row=row_num, column=4).value = papel.aseveracion
        ws.cell(row=row_num, column=5).value = papel.importancia
        ws.cell(row=row_num, column=6).value = papel.obligatorio
        ws.cell(row=row_num, column=7).value = papel.descripcion or ""

        # Aplicar estilos a datos
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Congelar header
    ws.freeze_panes = "A2"

    # Convertir a BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("/{cliente_id}/plantilla")
async def descargar_plantilla(
    cliente_id: str,
    ls: Optional[int] = Query(None, description="Línea de Cuenta (ej: 130, 140) - opcional"),
    user: UserContext = Depends(get_current_user),
    session: Any = Depends(get_session),
):
    """
    Descargar plantilla Excel para papeles de trabajo

    Parámetros:
    - cliente_id: ID del cliente
    - ls (opcional): Línea de Cuenta (ej: 130) para descargar solo esos papeles

    Retorna: Excel con estructura de papeles de trabajo
    """
    try:
        # Obtener papeles
        query = session.query(WorkpapersTemplate)

        if ls is not None:
            query = query.filter(WorkpapersTemplate.ls == ls)

        papeles = query.order_by(WorkpapersTemplate.ls, WorkpapersTemplate.numero).all()

        if not papeles:
            message = f"No se encontraron papeles de trabajo para L/S {ls}" if ls else "No se encontraron papeles de trabajo"
            raise_api_error(
                status_code=status.HTTP_404_NOT_FOUND,
                code="NO_PAPELES_FOUND",
                message=message,
            )

        # Crear workbook
        excel_file = _create_template_workbook(papeles)

        # Determinar nombre de archivo
        if ls:
            filename = f"plantilla_papeles_trabajo_LS{ls}.xlsx"
        else:
            filename = f"plantilla_papeles_trabajo_completa.xlsx"

        return StreamingResponse(
            iter([excel_file.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        # Si es una excepción de API ya manejada, dejarla pasar
        if hasattr(e, "status_code"):
            raise
        # Otros errores
        raise_api_error(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            code="TEMPLATE_GENERATION_ERROR",
            message=f"Error generando plantilla: {str(e)}",
        )
