"""
Endpoints para upload y lectura de Trial Balance y Libro Mayor
"""
from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, File, UploadFile, status
from fastapi.responses import JSONResponse

from backend.auth import authorize_cliente_access, get_current_user
from backend.schemas import ApiResponse, UserContext
from backend.utils.api_errors import raise_api_error

LOGGER = logging.getLogger("socio_ai.trial_balance")

router = APIRouter(prefix="/api/trial-balance", tags=["trial-balance"])

ROOT = Path(__file__).resolve().parents[2]
DATA_CLIENTES = ROOT / "data" / "clientes"

ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv"}
MAX_FILE_SIZE_MB = 20


def _cliente_dir(cliente_id: str) -> Path:
    return DATA_CLIENTES / cliente_id


def _validate_file(file: UploadFile, kind: str) -> None:
    filename = file.filename or ""
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise_api_error(
            status_code=status.HTTP_400_BAD_REQUEST,
            code="INVALID_FILE_TYPE",
            message=f"Tipo de archivo no permitido: {ext}. Use .xlsx, .xls o .csv",
        )


@router.post("/{cliente_id}/upload", response_model=ApiResponse)
async def upload_trial_balance(
    cliente_id: str,
    file: UploadFile = File(...),
    kind: str = "tb",
    user: UserContext = Depends(get_current_user),
) -> ApiResponse:
    """
    Subir Trial Balance (kind=tb) o Libro Mayor (kind=mayor) para un cliente.
    El archivo se guarda en data/clientes/{cliente_id}/tb.xlsx o mayor.xlsx
    """
    authorize_cliente_access(cliente_id, user)

    if kind not in {"tb", "mayor"}:
        raise_api_error(
            status_code=status.HTTP_400_BAD_REQUEST,
            code="INVALID_KIND",
            message="El parametro kind debe ser 'tb' o 'mayor'",
        )

    _validate_file(file, kind)

    # Leer contenido del archivo
    content = await file.read()
    size_mb = len(content) / (1024 * 1024)
    if size_mb > MAX_FILE_SIZE_MB:
        raise_api_error(
            status_code=status.HTTP_400_BAD_REQUEST,
            code="FILE_TOO_LARGE",
            message=f"El archivo excede el limite de {MAX_FILE_SIZE_MB}MB",
        )

    # Crear directorio del cliente si no existe
    cliente_dir = _cliente_dir(cliente_id)
    cliente_dir.mkdir(parents=True, exist_ok=True)

    # Guardar siempre como .xlsx o .csv segun extension original
    original_ext = Path(file.filename or "archivo.xlsx").suffix.lower()
    # Normalizar: siempre guardar como tb.xlsx o mayor.xlsx para compatibilidad
    target_filename = f"{kind}.xlsx" if original_ext in {".xlsx", ".xls"} else f"{kind}.csv"
    target_path = cliente_dir / target_filename

    try:
        with open(target_path, "wb") as f:
            f.write(content)
    except Exception as e:
        raise_api_error(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            code="FILE_SAVE_ERROR",
            message=f"Error al guardar el archivo: {str(e)}",
        )

    return ApiResponse(
        data={
            "stored_as": target_filename,
            "original_name": file.filename or target_filename,
            "size_bytes": len(content),
            "kind": kind,
            "cliente_id": cliente_id,
        }
    )


@router.get("/{cliente_id}/diagnostico", response_model=ApiResponse)
def get_tb_diagnostico(
    cliente_id: str,
    user: UserContext = Depends(get_current_user),
) -> ApiResponse:
    """
    Diagnóstico del TB cargado: columnas detectadas, filas, stage
    """
    authorize_cliente_access(cliente_id, user)
    try:
        from analysis.lector_tb import obtener_diagnostico_tb
        diag = obtener_diagnostico_tb(cliente_id)
        return ApiResponse(data=diag)
    except Exception as e:
        return ApiResponse(data={"error": str(e), "cliente_id": cliente_id})


@router.get("/{cliente_id}/status", response_model=ApiResponse)
def get_tb_status(
    cliente_id: str,
    user: UserContext = Depends(get_current_user),
) -> ApiResponse:
    """
    Verificar si el cliente tiene TB y Mayor cargados
    """
    authorize_cliente_access(cliente_id, user)

    cliente_dir = _cliente_dir(cliente_id)
    tb_xlsx_path = cliente_dir / "tb.xlsx"
    tb_csv_path = cliente_dir / "tb.csv"
    mayor_xlsx_path = cliente_dir / "mayor.xlsx"
    mayor_csv_path = cliente_dir / "mayor.csv"

    tb_path = tb_xlsx_path if tb_xlsx_path.exists() else tb_csv_path
    mayor_path = mayor_xlsx_path if mayor_xlsx_path.exists() else mayor_csv_path

    has_tb = tb_path.exists() and tb_path.stat().st_size > 0
    has_mayor = mayor_path.exists() and mayor_path.stat().st_size > 0

    return ApiResponse(
        data={
            "cliente_id": cliente_id,
            "has_tb": has_tb,
            "has_mayor": has_mayor,
            "tb_size_bytes": tb_path.stat().st_size if has_tb else 0,
            "mayor_size_bytes": mayor_path.stat().st_size if has_mayor else 0,
        }
    )


@router.get("/{cliente_id}/data-simple", response_model=ApiResponse)
def get_tb_data_simple(
    cliente_id: str,
    user: UserContext = Depends(get_current_user),
) -> ApiResponse:
    """
    Obtener datos REALES del Trial Balance del cliente
    """
    LOGGER.info(f"📂 Solicitud de datos TB para cliente: {cliente_id}")
    authorize_cliente_access(cliente_id, user)

    try:
        import openpyxl

        cliente_dir = _cliente_dir(cliente_id)
        LOGGER.debug(f"Directorio del cliente: {cliente_dir}")

        # Buscar archivo tb.xlsx
        tb_xlsx_path = cliente_dir / "tb.xlsx"
        tb_csv_path = cliente_dir / "tb.csv"

        tb_path = None
        if tb_xlsx_path.exists():
            tb_path = tb_xlsx_path
            LOGGER.info(f"✅ Archivo Excel encontrado: {tb_xlsx_path}")
        elif tb_csv_path.exists():
            tb_path = tb_csv_path
            LOGGER.info(f"✅ Archivo CSV encontrado: {tb_csv_path}")

        if not tb_path or not tb_path.exists():
            # Fallback a datos ficticios si no existe archivo
            LOGGER.warning(f"⚠️  No trial balance file found for {cliente_id}, usando datos de prueba")
            test_data = {
                "140 - Activos Intangibles": 150000,
                "170 - Propiedad Planta Equipo": 500000,
                "130 - Cuentas por Cobrar": 250000,
                "210 - Cuentas por Pagar": 100000,
                "400 - Ingresos por Ventas": 1000000,
                "410 - Gastos de Operación": 600000,
            }
            return ApiResponse(
                data={
                    "cliente_id": cliente_id,
                    "datos": test_data,
                    "mensaje": "Datos de prueba (no se encontró archivo real)",
                }
            )

        # Leer archivo Excel real
        LOGGER.debug(f"🔖 Leyendo archivo: {tb_path}")
        data_dict: dict[str, Any] = {}
        wb = openpyxl.load_workbook(tb_path)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        LOGGER.debug(f"Total de filas en archivo: {len(rows)}")

        if len(rows) < 2:
            LOGGER.warning(f"⚠️  Archivo vacío para {cliente_id}")
            return ApiResponse(data={"error": "Archivo vacío", "cliente_id": cliente_id})

        # Primera fila es header
        headers = [str(h).lower() if h else "" for h in rows[0]]
        LOGGER.debug(f"Encabezados detectados: {headers}")

        # Encontrar índices de cuenta y saldo
        cuenta_idx = None
        saldo_idx = None

        for i, h in enumerate(headers):
            if 'cuenta' in h or 'nombre' in h or 'descripcion' in h:
                cuenta_idx = i
            if 'saldo' in h or 'monto' in h or 'valor' in h or 'balance' in h:
                saldo_idx = i

        if cuenta_idx is None:
            cuenta_idx = 0
        if saldo_idx is None:
            saldo_idx = 1 if len(headers) > 1 else 0

        LOGGER.debug(f"Índices detectados - Cuenta: {cuenta_idx}, Saldo: {saldo_idx}")

        # Procesar datos
        processed_count = 0
        total_saldo = 0
        for row in rows[1:]:
            if len(row) > max(cuenta_idx, saldo_idx):
                cuenta = str(row[cuenta_idx]).strip() if row[cuenta_idx] else None
                try:
                    saldo = float(row[saldo_idx]) if row[saldo_idx] else 0
                except:
                    saldo = 0

                if cuenta and cuenta.lower() != 'none' and saldo != 0:
                    data_dict[cuenta] = saldo
                    processed_count += 1
                    total_saldo += saldo

        LOGGER.info(f"✅ Datos cargados: {processed_count} cuentas, Total balance: ${total_saldo:,.2f}")

        return ApiResponse(
            data={
                "cliente_id": cliente_id,
                "datos": data_dict,
                "total_cuentas": len(data_dict),
                "mensaje": "Datos reales del cliente",
            }
        )

    except Exception as e:
        LOGGER.error(f"❌ Error al leer TB para {cliente_id}: {e}", exc_info=True)
        # Fallback a datos ficticios si hay error
        test_data = {
            "140 - Activos Intangibles": 150000,
            "170 - Propiedad Planta Equipo": 500000,
            "130 - Cuentas por Cobrar": 250000,
            "210 - Cuentas por Pagar": 100000,
            "400 - Ingresos por Ventas": 1000000,
            "410 - Gastos de Operación": 600000,
        }
        return ApiResponse(
            data={
                "cliente_id": cliente_id,
                "datos": test_data,
                "error": str(e),
                "mensaje": "Datos de prueba (error al leer reales)",
            }
        )


@router.get("/{cliente_id}/data", response_model=ApiResponse)
def get_tb_data(
    cliente_id: str,
    user: UserContext = Depends(get_current_user),
) -> ApiResponse:
    """
    Obtener datos del Trial Balance en formato JSON
    """
    authorize_cliente_access(cliente_id, user)

    try:
        import openpyxl
        import csv

        cliente_dir = _cliente_dir(cliente_id)

        # Intentar leer archivo
        tb_xlsx_path = cliente_dir / "tb.xlsx"
        tb_csv_path = cliente_dir / "tb.csv"

        tb_path = None
        if tb_xlsx_path.exists():
            tb_path = tb_xlsx_path
        elif tb_csv_path.exists():
            tb_path = tb_csv_path

        if not tb_path or not tb_path.exists():
            raise_api_error(
                status_code=status.HTTP_404_NOT_FOUND,
                code="TB_NOT_FOUND",
                message=f"Trial Balance no cargado para cliente {cliente_id}",
            )

        # Leer archivo
        data_dict: dict[str, Any] = {}

        if str(tb_path).endswith(".xlsx"):
            # Leer Excel con openpyxl
            wb = openpyxl.load_workbook(tb_path)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return ApiResponse(data={"error": "Archivo vacío o sin datos", "cliente_id": cliente_id})

            # Primera fila es header
            headers = [str(h).lower() if h else "" for h in rows[0]]

            # Encontrar índices de cuenta y saldo
            cuenta_idx = None
            saldo_idx = None

            for i, h in enumerate(headers):
                if 'cuenta' in h or 'nombre' in h:
                    cuenta_idx = i
                if 'saldo' in h or 'monto' in h or 'valor' in h:
                    saldo_idx = i

            # Si no encuentra, usa primeras dos columnas
            if cuenta_idx is None:
                cuenta_idx = 0
            if saldo_idx is None:
                saldo_idx = 1 if len(headers) > 1 else 0

            # Procesar datos
            for row in rows[1:]:
                if len(row) > max(cuenta_idx, saldo_idx):
                    cuenta = str(row[cuenta_idx]).strip() if row[cuenta_idx] else None
                    try:
                        saldo = float(row[saldo_idx]) if row[saldo_idx] else 0
                    except:
                        saldo = 0

                    if cuenta and cuenta.lower() != 'none':
                        data_dict[cuenta] = saldo
        else:
            # Leer CSV
            with open(tb_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Buscar cuenta y saldo en columnas
                    cuenta = None
                    saldo = None

                    for key, val in row.items():
                        key_lower = key.lower() if key else ""
                        if 'cuenta' in key_lower or 'nombre' in key_lower:
                            cuenta = str(val).strip() if val else None
                        if 'saldo' in key_lower or 'monto' in key_lower or 'valor' in key_lower:
                            try:
                                saldo = float(val) if val else 0
                            except:
                                saldo = 0

                    if cuenta and saldo:
                        data_dict[cuenta] = saldo

        return ApiResponse(
            data={
                "cliente_id": cliente_id,
                "datos": data_dict,
                "total_cuentas": len(data_dict),
            }
        )

    except Exception as e:
        import traceback
        return ApiResponse(
            data={
                "error": str(e),
                "traceback": traceback.format_exc(),
                "cliente_id": cliente_id,
                "message": "Error al leer trial balance",
            }
        )
